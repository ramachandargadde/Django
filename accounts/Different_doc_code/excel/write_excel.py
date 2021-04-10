from openpyxl import Workbook
from openpyxl.styles import PatternFill
wb=Workbook()
wb.active
print(wb.active.title)
print(wb.sheetnames)
#set tile to sheet
wb['Sheet'].title="Report1"
sh1=wb.active
sh1['A1'].value='Name'
sh1['B1'].value="Status"
sh1['A2'].value='Name'
sh1['A2'].fill=PatternFill("solid",fgColor="71FF33")
sh1['B2'].value="Status"
sh1['B2'].fill=PatternFill("solid",fgColor="F50707")
wb.save("Final_report.xlsx")

