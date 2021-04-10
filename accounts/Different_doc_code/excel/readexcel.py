import  openpyxl

wb=openpyxl.load_workbook("excel/Book1.xlsx")
#openpyxl.load_workbook("c:\\users\\User\\Book1.xlsx")
print(type(wb))
sheets=wb.sheetnames
print(sheets)
print(wb.active.title)
sh1=wb["Sheet1"]
print(type(sh1))
data=sh1['b2'].value
print(data)
print(wb["Sheet1"]['A2'].value)
print(sh1.cell(3,2).value )
# rows = sh1.max_rows
# print(rows)
row = sh1.max_row
col=sh1.max_column
print(row)
print(col)
for i in range(1,row+1):
    for j in range(1,col+1):
        print(sh1.cell(i,j).value)


sh1.cell(row+1,1,value='3')
sh1.cell(row+1,2,value='ramesh')
sh1.cell(row+1,3,value='123@gmail.com')
wb.save('Book2.xlsx')