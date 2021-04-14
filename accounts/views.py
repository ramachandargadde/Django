from django.shortcuts import render, redirect
from django.http import HttpResponse
from django.forms import inlineformset_factory
from django.contrib.auth.forms import UserCreationForm

from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.models import User
from django.contrib import messages

from django.contrib.auth.decorators import login_required

# Create your views here.
from .models import *
from .forms import OrderForm, CreateUserForm,CustomerForm
from .filters import OrderFilter
from .decorators import unauthenticated_user, allowed_users, admin_only

from django.contrib.auth.models import Group
import csv,io
# from weasyprint import HTML
# import tempfile
import xlwt
from openpyxl import Workbook,load_workbook
from django.views.generic import View



@unauthenticated_user
def register(request):
    form = CreateUserForm()
    if request.method == 'POST':
        form = CreateUserForm(request.POST)
        if form.is_valid():
            user = form.save()
            username = form.cleaned_data.get('username')

            group = Group.objects.get(name='customer')
            user.groups.add(group)
            Customer.objects.create(
                user=user,
                name=user.username,
            )

            messages.success(request, 'Account was created for ' + username)

            return redirect('login')

    context = {'form': form}
    return render(request, 'accounts/register.html', context)


@unauthenticated_user
def loginpage(request):
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')

        user = authenticate(request, username=username, password=password)

        if user is not None:
            login(request, user)
            return redirect('home')
        else:
            messages.info(request, 'Username OR password is incorrect')

    context = {}
    return render(request, 'accounts/login.html', context)

def logoutUser(request):
	logout(request)
	return redirect('login')

@login_required(login_url='login')
@admin_only
def home(request):
    orders=Order.objects.all()
    customers=Customer.objects.all()
    total_customers=customers.count()
    total_orders=orders.count()
    delivered=orders.filter(status='Delivered').count()
    pending = orders.filter(status='Pending').count()
    custom={'orders':orders,
            'customers':customers,
            'total_orders':total_orders,
            'pending':pending,
            'delivered':delivered}
    return render(request, 'accounts/dashboard.html',custom)
@login_required(login_url='login')
@allowed_users(allowed_roles=['admin'])
def customer(request,pk_test):
    customer = Customer.objects.get(id=pk_test)
    orders = customer.order_set.all()
    order_count = orders.count()
    myFilter=OrderFilter(request.GET,queryset=orders)
    orders =myFilter.qs
    context = {'customer': customer, 'orders': orders, 'order_count': order_count,'myFilter':myFilter}
    return render(request, 'accounts/customer.html',context)

@login_required(login_url='login')
@allowed_users(allowed_roles=['customer'])
def userPage(request):
	orders = request.user.customer.order_set.all()

	total_orders = orders.count()
	delivered = orders.filter(status='Delivered').count()
	pending = orders.filter(status='Pending').count()

	print('ORDERS:', orders)

	context = {'orders':orders, 'total_orders':total_orders,
	'delivered':delivered,'pending':pending}
	return render(request, 'accounts/user.html', context)

@login_required(login_url='login')
@allowed_users(allowed_roles=['customer'])
def accountSettings(request):
	customer = request.user.customer
	form = CustomerForm(instance=customer)
	if request.method == 'POST':
		form = CustomerForm(request.POST, request.FILES,instance=customer)
		if form.is_valid():
			form.save()
	context = {'form':form}
	return render(request, 'accounts/account_settings.html', context)


@login_required(login_url='login')
@allowed_users(allowed_roles=['admin'])
def product(request):
    products = Product.objects.all()
    return render(request, 'accounts/product.html', {'products':products})

@login_required(login_url='login')
@allowed_users(allowed_roles=['admin'])
def createOrder(request, pk):
	OrderFormSet = inlineformset_factory(Customer, Order, fields=('product', 'status'), extra=10 )
	customer = Customer.objects.get(id=pk)
	formset = OrderFormSet(queryset=Order.objects.none(),instance=customer)
	#form = OrderForm(initial={'customer':customer})
	if request.method == 'POST':
		#print('Printing POST:', request.POST)
		#form = OrderForm(request.POST)
		formset = OrderFormSet(request.POST, instance=customer)
		if formset.is_valid():
			formset.save()
			return redirect('/')

	context = {'form':formset}
	return render(request, 'accounts/order_form.html', context)

@login_required(login_url='login')
@allowed_users(allowed_roles=['admin'])
def updateOrder(request,pk):
    order=Order.objects.get(id=pk)
    form = OrderForm(instance=order)
    if request.method=='POST':
        form = OrderForm(request.POST,instance=order)
        if form.is_valid():
            form.save()
            return redirect('/')
    context={'form':form}
    return render(request,'accounts/order_form.html',context)
@login_required(login_url='login')
@allowed_users(allowed_roles=['admin'])
def deleteOrder(request, pk):
	order = Order.objects.get(id=pk)
	if request.method == "POST":
		order.delete()
		return redirect('/')
	context = {'item':order}
	return render(request, 'accounts/delete.html', context)

@login_required(login_url='login')
@allowed_users(allowed_roles=['admin'])
def remove(request):
    print(request.POST)
    return redirect('/')



def export_csv(request):
    responce=HttpResponse(content_type='text/csv')
    responce['Content-Disposition']='attachment ; filename=Export_Product_details.csv'
    writer=csv.writer(responce)
    writer.writerow(['Product','Note','Date','Status'])
    orders=request.user.customer.order_set.all()
    for order in orders:
        writer.writerow([order.product,order.note,order.date_created,order.status])
    return responce
def export_excel(request):
    responce=HttpResponse(content_type='application/ms-excel')
    responce['Content-Disposition'] = 'attachment ; filename=ExportProductdetails.xlsx'
    wb=Workbook()
    wb['Sheet'].title = "Export_Product_details"
    sh1 = wb.active
    print(sh1)

    orders = request.user.customer.order_set.all()
    print(orders)
    i=1

    for order in orders:
        print(order.product, order.note, order.date_created, order.status)
        print(order.product)
        sh1.cell(1, 1).value =str(order.product)
        print(str(order.product))
        # sh1.cell(i, 3).value = order.note
        # sh1.cell(i, 4).value = order.date_created
        # sh1.cell(i, 5).value = order.status
        # print(order.product,order.note,order.date_created,order.status)
        # i=i+1

    wb.save("ExportProductdetails.xlsx")
    return responce

