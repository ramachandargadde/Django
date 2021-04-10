from openpyxl import load_workbook
from datetime import timedelta

wb=load_workbook("Demo_excel.xlsx") #excel name
sh1=wb["Duration_calculator"]  #sheet name in excel
for i in range(2,sh1.max_row+1): #edit rows and column numbers as required
    startdate=sh1.cell(i,1).value
    enddate=sh1.cell(i,3).value
    duration=int(enddate[0])*7
    final_date=str(startdate + timedelta(days=duration))
    sh1.cell(i, 2).value="URC-"+final_date.split()[0]
    wb.save("Demo_excel.xlsx")

    # print(duration)
    # print(startdate)
    # print(type(startdate))
    # print(startdate+timedelta(days=duration))
    # cha=startdate.split()
    # print(wb.sheetnames)
    # from datetime import datetime